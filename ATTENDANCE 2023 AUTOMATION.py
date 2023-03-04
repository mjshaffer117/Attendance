# ATTENDANCE 2023 AUTOMATION

import os.path, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Color, colors, PatternFill

DIR = "C:\\Users\\MJSha\\OneDrive\\Desktop\\Attendance Automation (Sarika)"
extraction_file = DIR + "\\" + "CSV-Template-File.csv"
# The file path and name where the information will be pulled from

def getDate(data):
    str_date = data[0][0]
    # Date & Time location in CSV - header columns

    month_list = {
    1 : 'Jan',
    2 : 'Feb',
    3 : 'Mar',
    4 : 'Apr',
    5 : 'May',
    6 : 'Jun',
    7 : 'Jul',
    8 : 'Aug',
    9 : 'Sep',
    10 : 'Oct',
    11 : 'Nov',
    12 : 'Dec'
    }

    month = int(str_date[:str_date.find('/')])
    day = str_date[str_date.find('/') + 1:str_date.find('/', str_date.find('/') + 1)]
    for key, value in month_list.items():
        if month == key:
            month = value
            date = month + " - " + day
            break
        else:
            date = 'NO VALUE'
            # Unknown value
    return date

def getQuestion(data):
    question_key = data[0][5]
    if question_key != '' and question_key != 'QUESTION':
        return question_key
    else:
        print("No value in question field.\n")
        print("Function cancelled.\n")
        input()
        return 0 # False
    
def getAnswerKey(data):
    answer_key = data[0][4]
    if answer_key != '' and answer_key != 'ANSWER':
        return answer_key
    else:
        print("No value in answer field.\n")
        print("Function cancelled.\n")
        input()
        return 0 # False

def getNames(data):
    names = []
    for i in range(len(data)):
        full_name = data[i][2] + " " + data[i][1]
        names.append(full_name)
        # Names are in lastname firstname order
    return names

def getUserAnswers(data):
    user_answers = []
    for i in range(len(data)):
        user_answers.append(data[i][3])
    return user_answers

def setDate(ws, date):
    i, j = 2, 2
    cell = ws.cell(row = i, column = j)

    alignment = Alignment(
        horizontal= "center",
        vertical= "bottom",
        text_rotation= 0,
        wrap_text= False,
        shrink_to_fit= False,
        indent= 0
    )

    while cell.value:
        #print(cell.value)
        j += 1
        cell = ws.cell(row = i, column = j)
    if not cell.value:
        cell = ws.cell(row = i, column = j, value = date).alignment = alignment

def setQuestion(ws, question):
    i, j = 3, 2
    cell = ws.cell(row = i, column = j)
    while cell.value:
        j += 1
        cell = ws.cell(row = i, column = j)
    if not cell.value:
        cell = ws.cell(row = i, column = j, value = question)
        return j

def checkAttendance(ws, answer_key, answers, names, start_column):
    i, j = 4, 1
    tmp = i
    cell = ws.cell(row = i, column = j)
    names_answers = {names: answers for names, answers in zip(names, answers)}
    #print(names_answers)

    set_fill = PatternFill(
        start_color= '7817B4',
        end_color= '7817B4',
        fill_type= 'solid'
    )

    while cell.value:
        if cell.value in names_answers:
            for name, answer in names_answers.items():
                if name == cell.value:
                    if answer == answer_key:
                        credit = 1
                    else:
                        credit = 0
                    ws.cell(row = i, column = start_column, value = credit)
                    break
            i += 1
            cell = ws.cell(row = i, column = j)
        else:
            ws.cell(row = i, column = start_column).fill = set_fill
            i += 1
            cell = ws.cell(row = i, column = j)

def updateExcelDoc(question, answer_key, date, names, answers):
    target_file = DIR + "\\" + "ESP 2023 ATTENDANCE-Template-File.xlsx"
    # The target path and name where extracted information will be placed
    if os.path.exists(target_file):
        wb = load_workbook(filename = target_file)
        ws = wb["Sheet1"]

        setDate(ws, date)
        start_column = setQuestion(ws, question)
        checkAttendance(ws, answer_key, answers, names, start_column)

        wb.save(target_file)
        print("\nAutomation completed successfully!")
        input()

    else:
        print("\nTarget file " + target_file + " was not found.")
        input()


'''MAIN CODE'''
if os.path.exists(extraction_file):
    data_rows = []
    with open(extraction_file, 'r') as csv_file:
        csvreader = csv.reader(csv_file)
        next(csvreader)
        for row in csvreader:
            data_rows.append(row)
    csv_file.close()
    question = getQuestion(data_rows)
    answer = getAnswerKey(data_rows)
    if question and answer:
        date = getDate(data_rows)
        name_list = getNames(data_rows)
        user_answers = getUserAnswers(data_rows)
        
        updateExcelDoc(question, answer, date, name_list, user_answers)

else:
    print("\nError finding CSV file.\n")
    print(extraction_file + "\nwas not found.")
    input()